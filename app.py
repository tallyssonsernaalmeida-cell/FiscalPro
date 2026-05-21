from flask import (Flask, render_template, request, redirect,
                   url_for, session, send_file, flash)
import pandas as pd
import os, json, io, re
from datetime import datetime, timedelta
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import ofxparse
from openpyxl import Workbook
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', os.urandom(24))
app.jinja_env.filters['enumerate'] = enumerate

app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=1)

USERS_FILE         = 'users.json'
SUPPORT_FILE       = 'support.json'
LOGS_FILE          = 'logs.json'
UPLOAD_FOLDER      = 'uploads'
USER_DATA          = 'user_data'
CONCILIACAO_FOLDER = 'conciliacao'

for d in [UPLOAD_FOLDER, USER_DATA, CONCILIACAO_FOLDER]:
    os.makedirs(d, exist_ok=True)

ALLOWED = {'ofx', 'xlsx', 'xls', 'xlsm', 'csv', 'pdf'}

# ─────────────────────────────────────────────────────────────
@app.after_request
def nao_cachear(response):
    if 'user' in session:
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        response.headers['Pragma']        = 'no-cache'
        response.headers['Expires']       = '0'
    return response

# ─────────────────────────────────────────────────────────────
# LOGS
# ─────────────────────────────────────────────────────────────
def add_log(acao, usuario='sistema', detalhe=''):
    logs = []
    if os.path.exists(LOGS_FILE):
        try:
            with open(LOGS_FILE, 'r', encoding='utf-8') as f:
                logs = json.load(f)
        except:
            logs = []
    ip = detalhe.replace('IP: ', '') if detalhe.startswith('IP:') else (detalhe or '-')
    logs.insert(0, {
        'data':    datetime.now().strftime('%d/%m/%Y %H:%M:%S'),
        'acao':    acao,
        'usuario': usuario,
        'detalhe': ip
    })
    logs_unicos = []
    for log in logs[:500]:
        if not logs_unicos or not (
            log['acao'] == logs_unicos[-1]['acao'] and
            log['usuario'] == logs_unicos[-1]['usuario'] and
            log['data'][:10] == logs_unicos[-1]['data'][:10]
        ):
            logs_unicos.append(log)
    with open(LOGS_FILE, 'w', encoding='utf-8') as f:
        json.dump(logs_unicos[:500], f, ensure_ascii=False, indent=2)

def get_logs():
    if not os.path.exists(LOGS_FILE):
        return []
    try:
        with open(LOGS_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except:
        return []

# ─────────────────────────────────────────────────────────────
# USUÁRIOS
# ─────────────────────────────────────────────────────────────
def load_users():
    if not os.path.exists(USERS_FILE):
        default = {
            "admin@fiscal.app": {
                "password": generate_password_hash("admin123"),
                "nome": "Administrador",
                "role": "admin",
                "files": [],
                "whatsapp": "",
                "created_at": datetime.now().isoformat(),
                "ativo": True
            }
        }
        save_users(default)
        return default
    with open(USERS_FILE, 'r', encoding='utf-8') as f:
        return json.load(f)

def save_users(u):
    with open(USERS_FILE, 'w', encoding='utf-8') as f:
        json.dump(u, f, ensure_ascii=False, indent=4)

def load_support():
    if not os.path.exists(SUPPORT_FILE):
        return []
    try:
        with open(SUPPORT_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except:
        return []

def save_support(m):
    with open(SUPPORT_FILE, 'w', encoding='utf-8') as f:
        json.dump(m, f, ensure_ascii=False, indent=4)

def get_ctx():
    users = load_users()
    user = users.get(session.get('user'), {})
    return {
        'nome':     user.get('nome', ''),
        'email':    session.get('user', ''),
        'is_admin': user.get('role') == 'admin'
    }

def require_login():
    if 'user' not in session:
        return redirect(url_for('login'))
    return None

def require_admin():
    redir = require_login()
    if redir:
        return redir
    if load_users().get(session['user'], {}).get('role') != 'admin':
        flash('Acesso restrito.', 'warning')
        return redirect(url_for('dashboard'))
    return None

# ─────────────────────────────────────────────────────────────
# CATEGORIZAÇÃO
# ─────────────────────────────────────────────────────────────
def categorizar(desc):
    d = str(desc).lower()
    if 'estorno' in d:   return 'Estorno'
    if 'reembolso' in d: return 'Reembolso'
    if 'pix receb' in d or 'transferência recebida' in d or 'transferencia recebida' in d:
        return 'Transferência Recebida'
    if 'pix enviad' in d or 'transferência enviada' in d or 'transferencia enviada' in d:
        return 'Transferência Enviada'
    if 'ted receb' in d: return 'TED Recebido'
    if 'ted enviad' in d: return 'TED Enviado'
    if 'depósito' in d or 'deposito' in d: return 'Depósito'
    if 'crédito em conta' in d or 'credito em conta' in d: return 'Crédito em Conta'
    if 'pagamento de fatura' in d:  return 'Pagamento Fatura'
    if 'pagamento de boleto' in d:  return 'Pagamento Boleto'
    if 'débito automático' in d or 'debito automatico' in d: return 'Débito Automático'
    if 'pagamento' in d: return 'Pagamento'
    if 'compra no débito' in d or 'compra no debito' in d or 'nupay' in d:
        if 'pombo' in d or 'pombo24' in d: return 'Delivery'
        if 'uber' in d or '99' in d or 'cabify' in d: return 'Transporte'
        if any(x in d for x in ['panificadora','padaria','cafe','café','pub','lanchonete','restaurante','pizz']):
            return 'Alimentação'
        if any(x in d for x in ['supermercad','doca','mercado','hortifruti']): return 'Supermercado'
        if any(x in d for x in ['posto','combustivel','combustível','gasolina']): return 'Combustível'
        if 'tabacaria' in d: return 'Tabacaria'
        if 'pmb joi' in d or 'prefeitura' in d or 'municipio' in d: return 'Impostos & Taxas'
        return 'Compra Débito'
    if 'compra no crédito' in d or 'compra no credito' in d: return 'Compra Crédito'
    if any(x in d for x in ['salário','salario','holerite','folha','vencimento']): return 'Salário/Renda'
    if any(x in d for x in ['freelance','comissão','comissao','honorário','honorario']): return 'Renda Extra'
    if any(x in d for x in ['aluguel','condomínio','condominio','iptu','financiamento','prestação','prestacao']):
        return 'Moradia'
    if any(x in d for x in ['supermercado','mercado','hortifruti','açougue','acougue','padaria',
                              'restaurante','ifood','lanchonete','alimentação','alimentacao','comida']):
        return 'Alimentação'
    if any(x in d for x in ['uber','cabify','ônibus','onibus','metrô','metro','gasolina',
                              'combustível','combustivel','posto','estacionamento','pedágio','pedagio','transporte']):
        return 'Transporte'
    if any(x in d for x in ['energia','água','agua','internet','telefone','celular','claro','vivo',
                              'tim','net','gás','gas','assinatura','streaming','mensalidade','celesc']):
        return 'Contas & Serviços'
    if any(x in d for x in ['farmácia','farmacia','drogaria','médico','medico','hospital',
                              'plano de saúde','unimed','exame','consulta','remédio','remedio']):
        return 'Saúde'
    if any(x in d for x in ['netflix','spotify','amazon','disney','hbo','cinema','teatro',
                              'ingresso','show','parque','viagem']):
        return 'Lazer'
    if any(x in d for x in ['curso','faculdade','escola','educação','educacao','livro','livraria']):
        return 'Educação'
    if any(x in d for x in ['roupa','calçado','calcado','moda','loja','zara','renner','riachuelo','shopping']):
        return 'Vestuário'
    if any(x in d for x in ['imposto','taxa','multa','darf','irpf','irpj','sefaz','municip','pmb joi','prefeitura']):
        return 'Impostos & Taxas'
    if any(x in d for x in ['investimento','ações','acao','fii','fundo','cdb','tesouro','bolsa']):
        return 'Investimentos'
    if any(x in d for x in ['seguro','previdência','previdencia','vida','auto']):
        return 'Seguros'
    if any(x in d for x in ['google','microsoft','office','adobe','dropbox','icloud']):
        return 'Assinaturas Digitais'
    return 'Outros'

# ─────────────────────────────────────────────────────────────
# PROCESSAMENTO DE ARQUIVOS (UNIVERSAL - TODOS OS BANCOS)
# ─────────────────────────────────────────────────────────────
def process_file(fp):
    """Processa qualquer arquivo: OFX (BB, Itaú, Bradesco, Santander, Inter, Nubank...), Excel, CSV"""
    ext = fp.rsplit('.', 1)[-1].lower()
    
    if ext == 'ofx':
        txs = []
        
        # ═══ MÉTODO 1: ofxparse (OFX padrão) ═══
        try:
            with open(fp, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
            
            # Corrige encoding problemático
            content = content.replace('encoding="ISO-8859-1"', 'encoding="UTF-8"')
            content = content.replace('encoding="US-ASCII"', 'encoding="UTF-8"')
            
            temp_path = fp + '.tmp'
            with open(temp_path, 'w', encoding='utf-8') as f:
                f.write(content)
            
            with open(temp_path, 'r', encoding='utf-8') as f:
                ofx = ofxparse.OfxParser.parse(f)
            
            for acc in ofx.accounts:
                for t in acc.statement.transactions:
                    txs.append({
                        'data': str(t.date)[:10],
                        'descricao': (t.memo or '')[:200],
                        'valor': float(t.amount),
                        'tipo': 'Crédito' if float(t.amount) > 0 else 'Débito',
                        'categoria': categorizar(t.memo or ''),
                        'pago': False
                    })
            
            if os.path.exists(temp_path):
                os.remove(temp_path)
            
            if txs:
                print(f'OFX processado (método 1): {len(txs)} transações')
                return txs
        except Exception as e:
            print(f'Método 1 falhou: {e}')
        
        # ═══ MÉTODO 2: XML direto (BB, Caixa, Sicoob, etc.) ═══
        try:
            import xml.etree.ElementTree as ET
            
            with open(fp, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
            
            content = re.sub(r'<\?OFX.*?\?>', '', content, flags=re.DOTALL)
            content = re.sub(r'</?OFX[^>]*>', '', content, flags=re.DOTALL)
            content = re.sub(r'</?SONRS[^>]*>.*?</SONRS>', '', content, flags=re.DOTALL)
            content = re.sub(r'</?SIGNONMSGS[^>]*>.*?</SIGNONMSGS>', '', content, flags=re.DOTALL)
            
            root = ET.fromstring(content)
            
            for trn in root.iter('STMTTRN'):
                try:
                    dt = trn.find('DTPOSTED') or trn.find('DTACCT')
                    val = trn.find('TRNAMT')
                    desc = trn.find('MEMO') or trn.find('NAME') or trn.find('PAYEE')
                    
                    data = 'N/A'
                    if dt is not None and dt.text:
                        raw = dt.text.strip()[:8]
                        if len(raw) == 8:
                            data = f'{raw[:4]}-{raw[4:6]}-{raw[6:8]}'
                    
                    valor = 0.0
                    if val is not None and val.text:
                        valor = float(val.text.replace(',', '.'))
                    
                    descricao = 'Sem descrição'
                    if desc is not None and desc.text:
                        descricao = desc.text.strip()[:200]
                    
                    txs.append({
                        'data': data,
                        'descricao': descricao,
                        'valor': valor,
                        'tipo': 'Crédito' if valor > 0 else 'Débito',
                        'categoria': categorizar(descricao),
                        'pago': False
                    })
                except:
                    continue
            
            if txs:
                print(f'OFX processado (método 2 - XML): {len(txs)} transações')
                return txs
        except Exception as e:
            print(f'Método 2 falhou: {e}')
        
        # ═══ MÉTODO 3: Tentar como CSV (alguns bancos exportam OFX como CSV) ═══
        try:
            df = pd.read_csv(fp, sep=None, engine='python', encoding='utf-8')
            txs = _process_dataframe(df)
            if txs:
                print(f'Arquivo processado como CSV: {len(txs)} transações')
                return txs
        except:
            pass
        
        return txs
    
    # ═══ EXCEL / CSV ═══
    try:
        if ext == 'csv':
            for enc in ['utf-8', 'latin-1', 'iso-8859-1']:
                try:
                    df = pd.read_csv(fp, sep=None, engine='python', encoding=enc)
                    break
                except:
                    continue
        elif ext in ['xlsx', 'xlsm']:
            df = pd.read_excel(fp, engine='openpyxl')
        else:
            df = pd.read_excel(fp, engine='xlrd')
        
        return _process_dataframe(df)
    except Exception as e:
        print(f'Erro Excel/CSV: {e}')
        return []


def _process_dataframe(df):
    """Processa DataFrame pandas para lista de transações"""
    df.columns = [str(c).strip().lower() for c in df.columns]
    
    col_d = next((c for c in df.columns if any(k in c for k in ['data', 'date', 'dt', 'dia'])), None)
    col_v = next((c for c in df.columns if any(k in c for k in ['valor', 'value', 'amount', 'montante', 'total'])), None)
    
    col_t = None
    for c in df.columns:
        cl = c.lower()
        if cl in ['descrição', 'descricao', 'historico', 'histórico', 'memo', 'complemento', 'detalhe', 'lançamento', 'lancamento']:
            col_t = c
            break
    if not col_t:
        for c in df.columns:
            cl = c.lower()
            if ('descri' in cl or 'histor' in cl or 'lanc' in cl or 'memo' in cl) and 'identif' not in cl:
                col_t = c
                break
    
    out = []
    for _, row in df.iterrows():
        v = 0.0
        if col_v:
            try:
                raw = str(row[col_v]).replace('R$', '').replace(' ', '').strip()
                if raw not in ('', 'nan', 'None'):
                    if ',' in raw and '.' in raw:
                        raw = raw.replace('.', '').replace(',', '.') if raw.rfind(',') > raw.rfind('.') else raw.replace(',', '')
                    elif ',' in raw:
                        raw = raw.replace(',', '.')
                    v = float(raw)
            except:
                v = 0.0
        
        data = 'N/A'
        if col_d:
            raw = str(row[col_d]).strip()[:10]
            if re.match(r'\d{2}/\d{2}/\d{4}', raw):
                p = raw.split('/')
                data = f'{p[2]}-{p[1]}-{p[0]}'
            elif re.match(r'\d{4}-\d{2}-\d{2}', raw):
                data = raw
            else:
                data = raw
        
        desc = str(row[col_t])[:200] if col_t and str(row[col_t]) != 'nan' else 'Sem descrição'
        
        out.append({
            'data': data,
            'descricao': desc,
            'valor': v,
            'tipo': 'Crédito' if v > 0 else 'Débito',
            'categoria': categorizar(desc),
            'pago': False
        })
    return out

# ─────────────────────────────────────────────────────────────
# DASHBOARD DATA
# ─────────────────────────────────────────────────────────────
def get_dash(email):
    users = load_users()
    files = users.get(email, {}).get('files', [])
    txs = []
    for fi in files:
        path = fi.get('path', '')
        if os.path.exists(path):
            txs.extend(process_file(path))
    
    rec = sum(t['valor'] for t in txs if t['valor'] > 0)
    desp = sum(abs(t['valor']) for t in txs if t['valor'] < 0)
    saldo = round(rec - desp, 2)
    
    cat_d, cat_r = {}, {}
    for t in txs:
        cat = t['categoria']
        if t['valor'] < 0:
            cat_d[cat] = round(cat_d.get(cat, 0) + abs(t['valor']), 2)
        else:
            cat_r[cat] = round(cat_r.get(cat, 0) + t['valor'], 2)
    
    return {
        'transacoes':       txs,
        'transactions':     txs,
        'total_receitas':   round(rec, 2),
        'total_creditos':   round(rec, 2),
        'total_despesas':   round(desp, 2),
        'total_debitos':    round(desp, 2),
        'saldo':            saldo,
        'status':           'Lucro' if saldo > 0 else ('Prejuízo' if saldo < 0 else 'Equilibrado'),
        'total_transacoes': len(txs),
        'qtd_transacoes':   len(txs),
        'categorias':       cat_d,
        'categorias_rec':   cat_r,
        'top_categorias':   sorted(cat_d.items(), key=lambda x: x[1], reverse=True)[:8],
        'percentual_lucro': round((saldo / rec) * 100, 1) if rec > 0 else 0
    }

# ─────────────────────────────────────────────────────────────
# EXCEL PROFISSIONAL
# ─────────────────────────────────────────────────────────────
def _brd():
    s = Side(style='thin', color='D0D0D0')
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(cor):
    return PatternFill(start_color=cor, end_color=cor, fill_type='solid')

def _fnt(cor='1E293B', bold=False, size=10):
    return Font(name='Calibri', color=cor, bold=bold, size=size)

def build_excel_conciliacao(data, recomendacoes=None):
    G, GL, R, RL = '22a060', 'E8F8EF', 'C0392B', 'FDEDEC'
    AZUL = '1D4ED8'
    wb = Workbook()
    brd = _brd()
    
    rec_list  = [t for t in data['transactions'] if t['valor'] > 0]
    desp_list = [t for t in data['transactions'] if t['valor'] < 0]
    tot_r = data['total_receitas']
    tot_d = data['total_despesas']
    saldo = data['saldo']
    pct = data['percentual_lucro']
    cats_d = sorted(data['categorias'].items(), key=lambda x: x[1], reverse=True)
    cats_r = sorted(data['categorias_rec'].items(), key=lambda x: x[1], reverse=True)

    ws = wb.active
    ws.title = '📊 Painel Resumo'
    ws.sheet_view.showGridLines = False
    
    ws.merge_cells('A1:J1')
    ws['A1'].value = 'FISCALPRO — RELATÓRIO DE CONCILIAÇÃO BANCÁRIA'
    ws['A1'].font = Font(name='Calibri', bold=True, size=18, color='FFFFFF')
    ws['A1'].fill = _fill(G)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40
    
    ws.merge_cells('A2:J2')
    ws['A2'].value = f'Gerado em {datetime.now().strftime("%d/%m/%Y às %H:%M")} | {data.get("qtd_transacoes", 0)} transações'
    ws['A2'].font = _fnt('6B7280', size=10)
    ws['A2'].fill = _fill('F4F6F8')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    
    kpis = [
        ('A', 'RECEITAS', f'R$ {tot_r:,.2f}', G, GL),
        ('D', 'DESPESAS', f'R$ {tot_d:,.2f}', R, RL),
        ('G', 'SALDO', f'R$ {saldo:,.2f}', G if saldo >= 0 else R, GL if saldo >= 0 else RL)
    ]
    for col, label, val, clr, bg in kpis:
        ws.merge_cells(f'{col}4:{chr(ord(col)+2)}6')
        ws[f'{col}4'].value = f'{label}\n{val}'
        ws[f'{col}4'].font = Font(name='Calibri', bold=True, size=14, color=clr)
        ws[f'{col}4'].fill = _fill(bg)
        ws[f'{col}4'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws[f'{col}4'].border = brd
    ws.row_dimensions[4].height = 55
    
    r = 8
    ws.merge_cells(f'A{r}:J{r}')
    ws[f'A{r}'].value = '📈 MÉTRICAS'
    ws[f'A{r}'].font = Font(name='Calibri', bold=True, size=13, color=G)
    r += 1
    metricas = [
        ('Total Transações', str(data.get('qtd_transacoes', 0))),
        ('Margem de Lucro', f'{pct}%'),
        ('Status', data.get('status', 'N/A')),
        ('Ticket Médio Rec.', f'R$ {(tot_r/len(rec_list)):,.2f}' if rec_list else 'R$ 0,00'),
        ('Ticket Médio Desp.', f'R$ {(tot_d/len(desp_list)):,.2f}' if desp_list else 'R$ 0,00'),
    ]
    for i, (m, v) in enumerate(metricas):
        c1 = ws.cell(row=r+i, column=1, value=m)
        c1.font = Font(name='Calibri', bold=True, size=11)
        c1.fill = _fill('F4F6F8')
        c1.border = brd
        ws.merge_cells(f'A{r+i}:C{r+i}')
        c2 = ws.cell(row=r+i, column=4, value=v)
        c2.font = Font(name='Calibri', bold=True, size=12, color=AZUL)
        c2.fill = _fill('F4F6F8')
        c2.border = brd
        ws.merge_cells(f'D{r+i}:F{r+i}')
    r += len(metricas) + 1
    
    ws.merge_cells(f'A{r}:F{r}')
    ws[f'A{r}'].value = '🔴 TOP DESPESAS'
    ws[f'A{r}'].font = Font(name='Calibri', bold=True, size=13, color=R)
    r += 1
    for i, h in enumerate(['Categoria', 'Total (R$)', '%', 'Qtd.', 'Média'], 1):
        c = ws.cell(row=r, column=i, value=h)
        c.font = _fnt('FFFFFF', True, 11)
        c.fill = _fill(R)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = brd
    cat_start = r + 1
    r += 1
    for i, (cat, val) in enumerate(cats_d[:10]):
        pct_c = round(val/tot_d*100, 1) if tot_d > 0 else 0
        qtd = sum(1 for t in desp_list if t['categoria'] == cat)
        media = round(val/qtd, 2) if qtd > 0 else 0
        bg = RL if i % 2 == 0 else 'FFFFFF'
        for ci, v in enumerate([cat, round(val, 2), f'{pct_c}%', qtd, f'R$ {media:,.2f}'], 1):
            c = ws.cell(row=r, column=ci, value=v)
            c.fill = _fill(bg)
            c.border = brd
            c.font = _fnt(R, True) if ci == 2 else _fnt()
            c.alignment = Alignment(horizontal='left' if ci == 1 else 'center', vertical='center')
        r += 1
    cat_end = r - 1
    if cats_d:
        bar1 = BarChart()
        bar1.type = 'col'
        bar1.title = 'Top Despesas'
        bar1.style = 10
        bar1.width = 18
        bar1.height = 12
        bar1.add_data(Reference(ws, min_col=2, min_row=cat_start-1, max_row=cat_end), titles_from_data=True)
        bar1.set_categories(Reference(ws, min_col=1, min_row=cat_start, max_row=cat_end))
        bar1.series[0].graphicalProperties.solidFill = R
        ws.add_chart(bar1, f'G{cat_start-2}')
    r = cat_end + 3
    
    ws.merge_cells(f'A{r}:F{r}')
    ws[f'A{r}'].value = '🟢 TOP RECEITAS'
    ws[f'A{r}'].font = Font(name='Calibri', bold=True, size=13, color=G)
    r += 1
    for i, h in enumerate(['Categoria', 'Total (R$)', '%', 'Qtd.', 'Média'], 1):
        c = ws.cell(row=r, column=i, value=h)
        c.font = _fnt('FFFFFF', True, 11)
        c.fill = _fill(G)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = brd
    rec_start = r + 1
    r += 1
    for i, (cat, val) in enumerate(cats_r[:10]):
        pct_c = round(val/tot_r*100, 1) if tot_r > 0 else 0
        qtd = sum(1 for t in rec_list if t['categoria'] == cat)
        media = round(val/qtd, 2) if qtd > 0 else 0
        bg = GL if i % 2 == 0 else 'FFFFFF'
        for ci, v in enumerate([cat, round(val, 2), f'{pct_c}%', qtd, f'R$ {media:,.2f}'], 1):
            c = ws.cell(row=r, column=ci, value=v)
            c.fill = _fill(bg)
            c.border = brd
            c.font = _fnt(G, True) if ci == 2 else _fnt()
            c.alignment = Alignment(horizontal='left' if ci == 1 else 'center', vertical='center')
        r += 1
    rec_end = r - 1
    if cats_r:
        pie = PieChart()
        pie.title = 'Distribuição de Receitas'
        pie.style = 10
        pie.width = 16
        pie.height = 12
        pie.add_data(Reference(ws, min_col=2, min_row=rec_start-1, max_row=rec_end), titles_from_data=True)
        pie.set_categories(Reference(ws, min_col=1, min_row=rec_start, max_row=rec_end))
        ws.add_chart(pie, f'G{rec_end+2}')
    r = rec_end + 20
    
    if recomendacoes:
        ws.merge_cells(f'A{r}:J{r}')
        ws[f'A{r}'].value = '💡 RECOMENDAÇÕES'
        ws[f'A{r}'].font = Font(name='Calibri', bold=True, size=14, color=AZUL)
        r += 1
        for rec in recomendacoes:
            ws.merge_cells(f'A{r}:J{r}')
            ws[f'A{r}'].value = rec
            ws[f'A{r}'].font = Font(name='Calibri', size=11, color='374151')
            ws[f'A{r}'].fill = _fill('EFF6FF' if '✅' in rec else 'FEF2F2' if '⚠️' in rec else 'FFFBEB')
            r += 1
    
    for col in 'ABCDEFGHIJ':
        ws.column_dimensions[col].width = 18

    ws2 = wb.create_sheet('💰 Receitas')
    ws2.sheet_view.showGridLines = False
    ws2.merge_cells('A1:F1')
    ws2['A1'].value = f'💰 RECEITAS — Total: R$ {tot_r:,.2f}'
    ws2['A1'].font = Font(name='Calibri', bold=True, size=14, color='FFFFFF')
    ws2['A1'].fill = _fill(G)
    ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
    for i, h in enumerate(['Data', 'Descrição', 'Categoria', 'Valor', '%', 'Status'], 1):
        c = ws2.cell(row=2, column=i, value=h)
        c.font = _fnt('FFFFFF', True, 11)
        c.fill = _fill(G)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = brd
    for j, t in enumerate(sorted(rec_list, key=lambda x: x['data'], reverse=True)):
        rr = j + 3
        pct_t = round(t['valor']/tot_r*100, 1) if tot_r > 0 else 0
        bg = GL if j % 2 == 0 else 'FFFFFF'
        for ci, v in enumerate([t['data'], t['descricao'][:60], t['categoria'], round(t['valor'], 2), f'{pct_t}%', '✅' if t.get('pago') else '⚠️'], 1):
            c = ws2.cell(row=rr, column=ci, value=v)
            c.fill = _fill(bg)
            c.border = brd
            c.alignment = Alignment(horizontal='left' if ci == 2 else 'center', vertical='center')
    for col, w in zip('ABCDEF', [14, 55, 22, 15, 14, 12]):
        ws2.column_dimensions[col].width = w

    ws3 = wb.create_sheet('📤 Despesas')
    ws3.sheet_view.showGridLines = False
    ws3.merge_cells('A1:F1')
    ws3['A1'].value = f'📤 DESPESAS — Total: R$ {tot_d:,.2f}'
    ws3['A1'].font = Font(name='Calibri', bold=True, size=14, color='FFFFFF')
    ws3['A1'].fill = _fill(R)
    ws3['A1'].alignment = Alignment(horizontal='center', vertical='center')
    for i, h in enumerate(['Data', 'Descrição', 'Categoria', 'Valor', '%', 'Status'], 1):
        c = ws3.cell(row=2, column=i, value=h)
        c.font = _fnt('FFFFFF', True, 11)
        c.fill = _fill(R)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = brd
    for j, t in enumerate(sorted(desp_list, key=lambda x: x['data'], reverse=True)):
        rr = j + 3
        pct_t = round(abs(t['valor'])/tot_d*100, 1) if tot_d > 0 else 0
        bg = RL if j % 2 == 0 else 'FFFFFF'
        for ci, v in enumerate([t['data'], t['descricao'][:60], t['categoria'], round(abs(t['valor']), 2), f'{pct_t}%', '✅' if t.get('pago') else '⚠️'], 1):
            c = ws3.cell(row=rr, column=ci, value=v)
            c.fill = _fill(bg)
            c.border = brd
            c.alignment = Alignment(horizontal='left' if ci == 2 else 'center', vertical='center')
    for col, w in zip('ABCDEF', [14, 55, 22, 15, 14, 12]):
        ws3.column_dimensions[col].width = w

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out

# ─────────────────────────────────────────────────────────────
# CONCILIAÇÃO
# ─────────────────────────────────────────────────────────────
def calcular_conciliacao(rec_list, desp_list):
    total = len(rec_list) + len(desp_list)
    if total == 0:
        return {'confirmados': 0, 'com_par': 0, 'sem_par': 0, 'percentual': 0, 'pares': [], 'total': 0}
    
    rec_pareados = set()
    desp_pareados = set()
    pares = []
    
    for ir, r in enumerate(rec_list):
        if ir in rec_pareados:
            continue
        if (r.get('categoria') or '').lower() not in ('estorno', 'reembolso'):
            continue
        for id_, d in enumerate(desp_list):
            if id_ in desp_pareados:
                continue
            if abs(r['valor'] - abs(d['valor'])) < 0.01 and r['data'] == d['data']:
                rec_pareados.add(ir)
                desp_pareados.add(id_)
                pares.append({
                    'tipo': '🔄 Estorno/Reembolso',
                    'data': r['data'],
                    'valor': r['valor'],
                    'desc_rec': r['descricao'][:50],
                    'desc_desp': d['descricao'][:50]
                })
                break
    
    for ir, r in enumerate(rec_list):
        if ir in rec_pareados:
            continue
        for id_, d in enumerate(desp_list):
            if id_ in desp_pareados:
                continue
            if abs(r['valor'] - abs(d['valor'])) < 0.01 and r['data'] == d['data'] and r['valor'] > 0:
                palavras_r = set(re.findall(r'[a-zA-ZÀ-ú]{4,}', r['descricao'].lower()))
                palavras_d = set(re.findall(r'[a-zA-ZÀ-ú]{4,}', d['descricao'].lower()))
                comuns = palavras_r & palavras_d - {'pelo', 'para', 'pix', 'transferencia', 'transferência',
                                                     'recebida', 'enviada', 'recebido', 'enviado', 'debito',
                                                     'crédito', 'credito', 'conta', 'compra'}
                if comuns:
                    rec_pareados.add(ir)
                    desp_pareados.add(id_)
                    pares.append({
                        'tipo': '🔗 Par Identificado',
                        'data': r['data'],
                        'valor': r['valor'],
                        'desc_rec': r['descricao'][:50],
                        'desc_desp': d['descricao'][:50]
                    })
                    break
    
    com_par = len(rec_pareados) + len(desp_pareados)
    sem_par = total - com_par
    return {
        'confirmados': total,
        'com_par': com_par,
        'sem_par': sem_par,
        'percentual': 100.0,
        'pares': pares,
        'total': total
    }

# ─────────────────────────────────────────────────────────────
# ROTAS — AUTH
# ─────────────────────────────────────────────────────────────
@app.route('/')
def index():
    if 'user' not in session:
        return redirect(url_for('login'))
    users = load_users()
    role = users.get(session['user'], {}).get('role')
    return redirect(url_for('admin_dashboard') if role == 'admin' else url_for('dashboard'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    session.clear()
    error = ''
    prefill = ''
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        senha = request.form.get('senha', '')
        users = load_users()
        if email in users and check_password_hash(users[email]['password'], senha):
            if not users[email].get('ativo', True):
                error = 'Conta desativada.'
            else:
                session['user'] = email
                add_log('Login', email, f'IP: {request.remote_addr}')
                flash(f'Bem-vindo(a), {users[email]["nome"]}!', 'success')
                return redirect(url_for('admin_dashboard') if users[email].get('role') == 'admin' else url_for('dashboard'))
        else:
            error = 'Email ou senha inválidos.'
            prefill = email
    return render_template('login.html', error=error, prefill_email=prefill)

@app.route('/cadastro', methods=['GET', 'POST'])
def cadastro():
    session.clear()
    error = ''
    if request.method == 'POST':
        nome = request.form.get('nome', '').strip()
        email = request.form.get('email', '').strip().lower()
        pwd = request.form.get('password', '')
        cpwd = request.form.get('confirm_password', '')
        if not all([nome, email, pwd]):
            error = 'Preencha todos os campos.'
        elif pwd != cpwd:
            error = 'Senhas não conferem.'
        elif len(pwd) < 6:
            error = 'Mínimo 6 caracteres.'
        else:
            users = load_users()
            if email in users:
                error = 'E-mail já cadastrado.'
            else:
                users[email] = {
                    'password': generate_password_hash(pwd),
                    'nome': nome,
                    'role': 'user',
                    'files': [],
                    'whatsapp': request.form.get('whatsapp', '').strip(),
                    'created_at': datetime.now().isoformat(),
                    'ativo': True
                }
                save_users(users)
                add_log('Novo cadastro', email, f'IP: {request.remote_addr}')
                session['user'] = email
                flash(f'Bem-vindo(a), {nome}!', 'success')
                return redirect(url_for('dashboard'))
    return render_template('cadastro.html', error=error)

@app.route('/logout')
def logout():
    add_log('Logout', session.get('user', ''), f'IP: {request.remote_addr}')
    session.clear()
    flash('Você saiu.', 'info')
    return redirect(url_for('login'))

# ─────────────────────────────────────────────────────────────
# ROTAS — ADMIN
# ─────────────────────────────────────────────────────────────
@app.route('/admin')
def admin_dashboard():
    redir = require_admin()
    if redir:
        return redir
    users = load_users()
    ctx = get_ctx()
    lista = []
    for em, d in users.items():
        lista.append({
            'email': em,
            'nome': d['nome'],
            'role': d['role'],
            'files_count': len(d.get('files', [])),
            'whatsapp': d.get('whatsapp', ''),
            'created_at': d.get('created_at', '')[:10],
            'ativo': d.get('ativo', True)
        })
    logs_unicos = []
    for log in get_logs():
        key = (log['usuario'], log['acao'])
        if key not in [(l['usuario'], l['acao']) for l in logs_unicos]:
            logs_unicos.append(log)
    return render_template('admin_dashboard.html', **ctx, active='admin',
                           usuarios=lista,
                           total_usuarios=len(users),
                           total_ativos=sum(1 for d in users.values() if d.get('ativo', True)),
                           support_messages=load_support(),
                           logs=logs_unicos[:30],
                           site_url=request.host_url.rstrip('/'))

@app.route('/admin/criar-usuario', methods=['POST'])
def admin_criar_usuario():
    redir = require_admin()
    if redir:
        return redir
    nome = request.form.get('nome', '').strip()
    email = request.form.get('email', '').strip().lower()
    pwd = request.form.get('password', '')
    role = request.form.get('role', 'user')
    users = load_users()
    if email in users:
        flash('E-mail já cadastrado!', 'danger')
    elif not all([nome, email, pwd]):
        flash('Preencha todos os campos!', 'danger')
    else:
        users[email] = {
            'password': generate_password_hash(pwd),
            'nome': nome,
            'role': role,
            'files': [],
            'whatsapp': '',
            'created_at': datetime.now().isoformat(),
            'ativo': True
        }
        save_users(users)
        add_log('Admin criou usuário', session['user'], f'{email} ({role})')
        flash(f'Usuário {nome} ({role}) criado!', 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/toggle-ativo', methods=['POST'])
def toggle_ativo():
    redir = require_admin()
    if redir:
        return redir
    email = request.form.get('email', '').strip().lower()
    users = load_users()
    if email in users and email != session['user']:
        users[email]['ativo'] = not users[email].get('ativo', True)
        save_users(users)
        flash(f'Usuário {email} {"ativado" if users[email]["ativo"] else "desativado"}.', 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/remover-usuario', methods=['POST'])
def remover_usuario():
    redir = require_admin()
    if redir:
        return redir
    email = request.form.get('email', '').strip().lower()
    users = load_users()
    if email == session['user']:
        flash('Não pode remover sua própria conta!', 'danger')
    elif email in users:
        for f in users[email].get('files', []):
            if os.path.exists(f.get('path', '')):
                os.remove(f['path'])
        del users[email]
        save_users(users)
        flash(f'Usuário {email} removido.', 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/configuracoes', methods=['GET', 'POST'])
def configuracoes():
    redir = require_admin()
    if redir:
        return redir
    users = load_users()
    ctx = get_ctx()
    error = ''
    success = ''
    if request.method == 'POST':
        acao = request.form.get('acao', '')
        if acao == 'alterar_email':
            novo = request.form.get('novo_email', '').strip().lower()
            if not novo:
                error = 'E-mail vazio.'
            elif novo in users and novo != session['user']:
                error = 'E-mail já em uso.'
            else:
                users[novo] = users.pop(session['user'])
                save_users(users)
                session['user'] = novo
                success = 'E-mail alterado!'
        elif acao == 'alterar_senha':
            sa = request.form.get('senha_atual', '')
            ns = request.form.get('nova_senha', '')
            cs = request.form.get('confirmar_senha', '')
            if not check_password_hash(users[session['user']]['password'], sa):
                error = 'Senha atual incorreta.'
            elif ns != cs:
                error = 'Senhas não conferem.'
            elif len(ns) < 6:
                error = 'Mínimo 6 caracteres.'
            else:
                users[session['user']]['password'] = generate_password_hash(ns)
                save_users(users)
                success = 'Senha alterada!'
    return render_template('Configurações.html', **ctx, active='admin_config',
                           admin_email=session['user'], error=error, success=success,
                           config={}, sys_python='Python 3', sys_platform=os.name)

# ─────────────────────────────────────────────────────────────
# ROTAS — USUÁRIO
# ─────────────────────────────────────────────────────────────
@app.route('/dashboard')
def dashboard():
    redir = require_login()
    if redir:
        return redir
    users = load_users()
    user = users.get(session['user'])
    ctx = get_ctx()
    dash = get_dash(session['user'])
    files = user.get('files', []) if user else []
    has_data = len(dash.get('transactions', [])) > 0
    
    pagina = request.args.get('pagina', 1, type=int)
    por_pagina = 50
    transacoes = list(dash.get('transactions', []))
    total_paginas = max(1, (len(transacoes) + por_pagina - 1) // por_pagina)
    inicio = (pagina - 1) * por_pagina
    dash['transactions'] = transacoes[inicio:inicio + por_pagina]
    dash['pagina'] = pagina
    dash['total_paginas'] = total_paginas
    
    return render_template('dashboard.html', **ctx, active='dashboard',
                           dash=dash, files=files, has_data=has_data)

@app.route('/conciliacao')
def conciliacao():
    redir = require_login()
    if redir:
        return redir
    ctx = get_ctx()
    dash = get_dash(session['user'])
    rec = [t for t in dash['transactions'] if t['valor'] > 0]
    desp = [t for t in dash['transactions'] if t['valor'] < 0]
    
    conc = calcular_conciliacao(rec, desp)
    
    recomendacoes = []
    if dash['saldo'] < 0:
        recomendacoes.append('⚠️ Seu saldo está negativo. Reveja suas despesas.')
    if dash['total_despesas'] > dash['total_receitas']:
        recomendacoes.append('📉 Despesas superam receitas. Crie um orçamento.')
    if conc['com_par'] > 0:
        recomendacoes.append(f'🔗 {conc["com_par"]} transações com par identificado.')
    recomendacoes.append(f'✅ {conc["confirmados"]} de {conc["total"]} lançamentos confirmados (100%).')
    if not recomendacoes:
        recomendacoes.append('✅ Suas finanças estão saudáveis!')

    return render_template('conciliacao.html', **ctx, active='conciliacao',
                           creditos=rec, debitos=desp,
                           total_creditos=sum(t['valor'] for t in rec),
                           total_debitos=sum(abs(t['valor']) for t in desp),
                           conciliados=conc['confirmados'],
                           total_itens=conc['total'],
                           percentual=conc['percentual'],
                           recomendacoes=recomendacoes)

@app.route('/upload', methods=['POST'])
def upload_file():
    redir = require_login()
    if redir:
        return redir
    file = request.files.get('file') or request.files.get('arquivo')
    if not file or file.filename == '':
        flash('Nenhum arquivo.', 'warning')
        return redirect(url_for('dashboard'))
    if '.' in file.filename and file.filename.rsplit('.', 1)[1].lower() in ALLOWED:
        fname = secure_filename(f"{session['user']}_{datetime.now().strftime('%Y%m%d%H%M%S')}_{file.filename}")
        fpath = os.path.join(UPLOAD_FOLDER, fname)
        file.save(fpath)
        users = load_users()
        if session['user'] not in users:
            users[session['user']] = {'files': []}
        users[session['user']].setdefault('files', []).append({
            'name': file.filename,
            'path': fpath,
            'type': file.filename.rsplit('.', 1)[1].lower(),
            'date': datetime.now().isoformat()
        })
        save_users(users)
        add_log('Upload', session['user'], file.filename)
        
        txs = process_file(fpath)
        if txs:
            flash(f'Arquivo "{file.filename}" importado com {len(txs)} transações!', 'success')
        else:
            flash(f'Arquivo "{file.filename}" salvo, mas nenhuma transação encontrada.', 'warning')
    else:
        flash('Formato não permitido.', 'danger')
    return redirect(url_for('dashboard'))

@app.route('/remover-arquivo', methods=['POST'])
def remover_arquivo():
    redir = require_login()
    if redir:
        return redir
    filename = request.form.get('filename', '')
    users = load_users()
    nova = []
    for f in users.get(session['user'], {}).get('files', []):
        if f['name'] == filename:
            if os.path.exists(f.get('path', '')):
                os.remove(f['path'])
        else:
            nova.append(f)
    if session['user'] in users:
        users[session['user']]['files'] = nova
    save_users(users)
    flash('Arquivo removido.', 'info')
    return redirect(url_for('dashboard'))

@app.route('/configuracoes_usuario', methods=['GET', 'POST'])
def configuracoes_usuario():
    redir = require_login()
    if redir:
        return redir
    users = load_users()
    user = users.get(session['user'], {})
    ctx = get_ctx()
    error = ''
    success = ''
    if request.method == 'POST':
        acao = request.form.get('acao', '')
        if acao == 'alterar_senha':
            sa = request.form.get('senha_atual', '')
            ns = request.form.get('nova_senha', '')
            cs = request.form.get('confirmar_senha', '')
            if not check_password_hash(users[session['user']]['password'], sa):
                error = 'Senha atual incorreta.'
            elif ns != cs:
                error = 'Senhas não conferem.'
            elif len(ns) < 6:
                error = 'Mínimo 6 caracteres.'
            else:
                users[session['user']]['password'] = generate_password_hash(ns)
                save_users(users)
                success = 'Senha alterada!'
        else:
            users[session['user']]['nome'] = request.form.get('nome', user.get('nome', '')).strip()
            users[session['user']]['whatsapp'] = request.form.get('whatsapp', '').strip()
            save_users(users)
            success = 'Dados atualizados!'
    return render_template('user_configuracoes.html', **ctx, active='config',
                           usuario=users.get(session['user'], {}),
                           error=error, success=success)

@app.route('/suporte', methods=['GET', 'POST'])
def suporte():
    redir = require_login()
    if redir:
        return redir
    ctx = get_ctx()
    if request.method == 'POST':
        msg = request.form.get('mensagem', '').strip()
        if msg:
            msgs = load_support()
            msgs.append({
                'de': session['user'],
                'nome': ctx['nome'],
                'mensagem': msg,
                'data': datetime.now().isoformat(),
                'lido': False
            })
            save_support(msgs)
            flash('Mensagem enviada!', 'success')
    return render_template('suporte.html', **ctx, active='suporte')

# ─────────────────────────────────────────────────────────────
# DOWNLOADS (100% LIBERADOS)
# ─────────────────────────────────────────────────────────────
@app.route('/download/excel')
def download_excel():
    redir = require_login()
    if redir:
        return redir
    dash = get_dash(session['user'])
    return send_file(build_excel_conciliacao(dash),
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name=f'FiscalPro_{datetime.now().strftime("%Y%m%d")}.xlsx')

@app.route('/download/csv')
def download_csv():
    redir = require_login()
    if redir:
        return redir
    df = pd.DataFrame(get_dash(session['user'])['transactions'])
    out = io.StringIO()
    df.to_csv(out, index=False)
    out.seek(0)
    return send_file(io.BytesIO(out.getvalue().encode('utf-8')),
                     mimetype='text/csv',
                     as_attachment=True,
                     download_name=f'FiscalPro_{datetime.now().strftime("%Y%m%d")}.csv')

@app.route('/download/conciliacao')
def download_conciliacao():
    redir = require_login()
    if redir:
        return redir
    dash = get_dash(session['user'])
    recs = ['✅ Conciliação 100% concluída!', '📊 Relatório gerado pelo FiscalPro']
    out = build_excel_conciliacao(dash, recomendacoes=recs)
    out.seek(0)
    return send_file(out,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name=f'Conciliacao_{datetime.now().strftime("%Y%m%d")}.xlsx')

# ─────────────────────────────────────────────────────────────
if __name__ == '__main__':
    print('\n' + '='*50)
    print('  FiscalPro iniciando...')
    print('  URL:   http://127.0.0.1:5000')
    print('  Admin: admin@fiscal.app / admin123')
    print('='*50 + '\n')
    app.run(debug=True, host='127.0.0.1', port=5000)
